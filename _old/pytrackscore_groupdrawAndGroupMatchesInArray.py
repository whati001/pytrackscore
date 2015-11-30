#!/bin/python

#Library to Track Value to Excel and read Info from Excel
from openpyxl import *
from openpyxl.styles import *
from openpyxl.cell import *

class pytrackscore():
	def __init__(self, *args, **kwargs): 
		#init function
		self.workbook_path = None
		#self.defineWorkbookPath('/home/andreas/Desktop/pythonExcel/Turnierplan.xlsx')		
		#print(self.getWorkbookPath())
		#print(self.getWorkbookName())
		#self.readWorksheetGroups('Groupsname_initialsetup','A1')
		self.writeWorksheetTracking('Generated_21092015_2227',gamemode='mixedgroup')

	def readWorksheetGroups(self, *args):
		for idx, arg in enumerate(args):
			if idx == 0:
				self.initialsheet_name = arg
			elif idx == 1:
				self.startfield = arg
			#print(arg)
		
		#Read Workbook
		self.workbook = load_workbook(filename = self.workbook_path) 		
		#print(self.initialsheet_name)
		#Read Workbook
		self.initialsheet = self.workbook[self.initialsheet_name]
		#Read Groups
		#if we got a startfield from outside or not
		if not hasattr(self, 'startfield'):
			self.startfield = 'A1'
			self.idx = 1
		else:
			#if we from outside something, set counter
			self.idx = int(self.startfield[1])

		print(self.idx)
		self.groups = [[]]
		self.pos01 = 0
		self.repeat = 'true'
		while self.repeat == 'true':
			self.fieldname = self.startfield[0]+str(self.idx)
			self.fieldvalue = str(self.initialsheet[self.fieldname].value)
			print('Value form Field '+self.fieldname+' : '+self.fieldvalue)
			self.idx += 1
			self.groups[self.pos01].append(self.fieldvalue)
			#if there are two Fields empty behind another, stop reading
			if self.fieldvalue == 'None':
				del self.groups[self.pos01][-1]
				if str(self.initialsheet[self.startfield[0]+str(self.idx)].value) == 'None':
					self.repeat = 'false'
				else:
					self.pos01 += 1
					self.groups.append([])

			print(self.groups)
		print('All Groups were saved to self.groups.')

	def writeWorksheetTracking(self, trackingworksheetname,**kwargs):
		#Check with Gamemode will used
		if kwargs['gamemode'] == 'mixedgroup':
			self.mixedenable = 'true'
		else:
			self.mixedenabel = 'false'
	
		'''self.trackingworksheetname = trackingworksheetname
		#print(self.workbook.get_sheet_names())
		if trackingworksheetname not in self.workbook.get_sheet_names():
			print('New worksheet was created with sheetname: '+str(trackingworksheetname))
			self.trackingsheet = self.workbook.create_sheet(title = trackingworksheetname)
		else:
			print('Existing worksheet used with sheetname: '+str(trackingworksheetname))
			self.trackingsheet = self.workbook[trackingworksheetname]
		#self.workbook.save(filename = self.workbook_path)

		#Initial Drawing
		#Groupcolors
		self.groupcolors = ['99FF99','FF9966','99CCFF','CC99FF','FFFF66']
		
		#self.trackingsheet['D1'] = 'Gruppenspiele'
		#self.trackingsheet.merge_cells('D1:G1')
		#self.trackingsheet['D1'].font = Font(bold=True)
		#self.trackingsheet['D1'].alignment = Alignment(horizontal='center')	
		#All Groups with there teams
		self.startfield_block = [ord('J')-64,'4']	
		for count01, group in enumerate(self.groups):
			#All teams in one Group
			#print(group)
			startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
			for count02, team in enumerate(group):
				#One Team
				#Teams vertical
				print('Write '+str(team)+' to '+str(startfield[0])+str(startfield[1]+count02))
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)] = team
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].font = Font(bold=True)
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#if the Groupname is written start with horicontal writting
				if count02 != 0:
					print('Write '+str(team)+' to '+chr(ord(startfield[0])+count02)+str(startfield[1]))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])] = team
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))

					#Write X to the SAME TEAM fields
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)] = 'X'
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].fill = PatternFill(patternType='solid', start_color='cccccc')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				if count01 == 0:
					#First group loop modify column width
					print('Change Column Dimensions to 20 for : '+str(chr(ord(startfield[0])+count02)))
					self.trackingsheet.column_dimensions[chr(ord(startfield[0])+count02)].width = 25
		'''
		self.groups = [['Gruppe A', 'Black Swanz', 'JVP Hofstätten', 'Kugelrund', 'Die Crew', 'SHT','06A','07A'], ['Gruppe B', 'Fireball Hafning', 'Jerich Dispo', 'Bum - Bum', 'Jerich Hasen', 'Blue Eyes Racing','06B','07B'], ['Gruppe C', 'der Böse Wolf und seine Geißlein', 'BELLAFFAIR.AT', 'HFC Ballerinas', "GH Baumi's Bubble Trouble", 'Ball Buzztards','06C','07C'], ['Gruppe D', 'Schlümpfe', 'FC Ambrosi', 'S-TEC Fighters', 'Jerich Hoolts Durch', 'Formation 88','06D','07D'],['Gruppe E', 'EESchlümpfe', 'EEEFC Ambrosi', 'EEEES-TEC Fighters', 'EEEEEJerich Hoolts Durch', 'EEEEFormation 88','EEEE06D','EEEEE07D']]
		#Draw GroupGames
		self.startfield_group = ['B','4']
		self.game_count = len(self.groups)*((len(self.groups[0])-1)*(len(self.groups[0])-2))/2
		print(len(self.groups))
		print(len(self.groups[0]))
		print('Total there are : %d' % (self.game_count))
		#Draw GroupMatches
		self.matches_logic = [[[1,2],[3,1],[2,3]],[[1,2],[3,4],[1,3],[2,4],[4,1],[3,2]],[[1,2],[3,4],[5,1],[2,3],[4,5],[1,3],[2,5],[4,1],[5,3],[2,4]],[[1,2],[3,4],[5,6],[1,4],[6,3],[2,5],[6,1],[3,5],[4,2],[1,5],[3,2],[6,4],[3,1],[5,3],[2,6]],[[1,2],[3,4],[5,6],[7,1],[2,3],[4,5],[6,7],[1,3],[2,4],[3,7],[5,2],[1,6],[7,4],[1,5],[2,6],[5,7],[1,3],[3,6],[2,7],[2,7],[5,3],[4,6]]]
		self.matches = []
		for idx, group in enumerate(self.groups):
			self.matches.append([])
			self.matches[idx].append(group[0])
	
		groupfinish = 0
		for game in range(1,int(self.game_count)+1):
			#for Mixxed Gameplay
			if 'mixgroup' == 'mixgroup':
				#Get the current game from the group, group number is variable groupfinish!!!
				groupgame = int(game-((len(self.groups))*groupfinish))
				print('Game %d' % groupgame)
				#Save Value to self.matches on position groupgame-1 => Groupnumber from self.groups on postion groupgame -1 => Groubnumer, and the posiontion from list self.matches_logic, there are the algorith hardcoded
				self.matches[groupgame-1].append(self.groups[groupgame-1][self.matches_logic[len(self.groups[0])-4][groupfinish][0]]+'_'+self.groups[groupgame-1][self.matches_logic[len(self.groups[0])-4][groupfinish][1]])
				#self.matches[groupgame-1].append(self.groups[groupgame-1][1]+'_'+self.groups[groupgame-1][2])
				if game % (len(self.groups)) == 0:
					groupfinish += 1
					print('Groupfinish %d' % groupfinish)
				print(self.matches)
		#if kwargs['gamemode'] == 'mixedgroup':
			
		#Save Workbook
		#self.workbook.save(filename = self.workbook_path)
	
	#Define workbookPath
	def defineWorkbookPath(self, workbook_path):
		self.workbook_path = workbook_path

	#Define get WorkbookPath
	def getWorkbookPath(self):
		return str(self.workbook_path)

	#Define get Workbookname
	def getWorkbookName(self):
		return str(self.workbook_path.split('/')[-1])

class_pytrackscore = pytrackscore()
