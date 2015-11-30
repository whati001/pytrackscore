#!/bin/python

#Library to Track Value to Excel and read Info from Excel
from openpyxl import *
from openpyxl.styles import *
from openpyxl.cell import *

class pytrackscore():
	def __init__(self, *args, **kwargs): 
		#init function
		self.workbook_path = None
		self.defineWorkbookPath('/Users/Andi/Desktop/Scoreboard/Turnierplan.xlsx')
		print(self.getWorkbookPath())
		print(self.getWorkbookName())
		self.readWorksheetGroups('Groupsname_initialsetup','A1')
		self.writeWorksheetInitial('Generated_21092015_2227','8:30','7','5',10,'8',gamemode='mixedgroup')
		self.getTeamNames(1)
		self.writeMatchValue(1,2,4)

	def readWorksheetGroups(self, *args):
		for idx, arg in enumerate(args):
			if idx == 0:
				self.initialsheet_name = arg
			elif idx == 1:
				self.startfield = arg
			#print(arg)
		
		#Read Workbook
		self.workbook = load_workbook(filename = self.workbook_path) 		
		#print(self.initialsheet_name)
		#Read Workbook
		self.initialsheet = self.workbook[self.initialsheet_name]
		#Read Groups
		#if we got a startfield from outside or not
		if not hasattr(self, 'startfield'):
			self.startfield = 'A1'
			self.idx = 1
		else:
			#if we from outside something, set counter
			self.idx = int(self.startfield[1])

		print(self.idx)
		self.groups = [[]]
		self.pos01 = 0
		self.repeat = 'true'
		while self.repeat == 'true':
			self.fieldname = self.startfield[0]+str(self.idx)
			self.fieldvalue = str(self.initialsheet[self.fieldname].value)
			print('Value form Field '+self.fieldname+' : '+self.fieldvalue)
			self.idx += 1
			self.groups[self.pos01].append(self.fieldvalue)
			#if there are two Fields empty behind another, stop reading
			if self.fieldvalue == 'None':
				del self.groups[self.pos01][-1]
				if str(self.initialsheet[self.startfield[0]+str(self.idx)].value) == 'None':
					self.repeat = 'false'
				else:
					self.pos01 += 1
					self.groups.append([])

		print('All Groups were saved to self.groups.')
		print(self.groups)

	def writeWorksheetInitial(self, trackingworksheetname,starttime,playtime,breaktime,playtime_crisscross,breaktime_crisscross,**kwargs):
		#Check with Gamemode will used
		
		
		if kwargs['gamemode'] == 'mixedgroup':
			self.mixedenable = 'true'
		else:
			self.mixedenable = 'false'
		
		self.trackingworksheetname = trackingworksheetname
		self.starttime = datetime.datetime.strptime(starttime, "%H:%M")
		self.playtime = playtime
		self.breaktime = breaktime
		self.playtime_crisscross = playtime_crisscross
		self.breaktime_crisscross = breaktime_crisscross
		#print(self.workbook.get_sheet_names())
		if trackingworksheetname not in self.workbook.get_sheet_names():
			print('New worksheet was created with sheetname: '+str(trackingworksheetname))
			self.trackingsheet = self.workbook.create_sheet(title = trackingworksheetname)
		else:
			print('Existing worksheet used with sheetname: '+str(trackingworksheetname))
			self.trackingsheet = self.workbook[trackingworksheetname]
		#self.workbook.save(filename = self.workbook_path)

		#Initial Drawing
		#Groupcolors
		self.groupcolors = ['99FF99','FF9966','99CCFF','CC99FF','FFFF66','99FF99','FF9966','99CCFF','CC99FF']
		
		#self.trackingsheet['D1'] = 'Gruppenspiele'
		#self.trackingsheet.merge_cells('D1:G1')
		#self.trackingsheet['D1'].font = Font(bold=True)
		#self.trackingsheet['D1'].alignment = Alignment(horizontal='center')	
		#All Groups with there teams
		self.startfield_block = [ord('J')-64,'4']	
		for count01, group in enumerate(self.groups):
			#All teams in one Group
			#print(group)
			startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
			for count02, team in enumerate(group):
				for count03 in range(len(self.groups[count01])):
					print("Fill : %s" % get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#One Team
				#Teams vertical
				print('Write '+str(team)+' to '+str(startfield[0])+str(startfield[1]+count02))
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)] = team
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].font = Font(bold=True)
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
				#if the Groupname is written start with horicontal writting
				if count02 != 0:
					print('Write '+str(team)+' to '+chr(ord(startfield[0])+count02)+str(startfield[1]))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])] = team
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))

					#Write X to the SAME TEAM fields
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)] = 'X'
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].fill = PatternFill(patternType='solid', start_color='cccccc')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				if count01 == 0:
					#First group loop modify column width
					print('Change Column Dimensions to 20 for : '+str(chr(ord(startfield[0])+count02)))
					self.trackingsheet.column_dimensions[chr(ord(startfield[0])+count02)].width = 25
		
		#self.groups = [['Gruppe A', 'Black Swanz', 'JVP Hofstätten', 'Kugelrund', 'Die Crew', 'SHT'], ['Gruppe B', 'Fireball Hafning', 'Jerich Dispo', 'Bum - Bum', 'Jerich Hasen', 'Blue Eyes Racing'], ['Gruppe C', 'der Böse Wolf und seine Geißlein', 'BELLAFFAIR.AT', 'HFC Ballerinas', "GH Baumi's Bubble Trouble", 'Ball Buzztards'], ['Gruppe D', 'Schlümpfe', 'FC Ambrosi', 'S-TEC Fighters', 'Jerich Hoolts Durch', 'Formation 88']]
		#Draw GroupGames
		self.startfield_group = [ord('B')-64,'4']
		self.game_count = len(self.groups)*((len(self.groups[0])-1)*(len(self.groups[0])-2))/2
		print(len(self.groups))
		print(len(self.groups[0]))
		print('Total there are : %d' % (self.game_count))
		#Draw GroupMatches
		self.matches_logic = [[[1,2],[2,1]],[[1,2],[3,1],[2,3]],[[1,2],[3,4],[1,3],[2,4],[4,1],[3,2]],[[1,2],[3,4],[5,1],[2,3],[4,5],[1,3],[2,5],[4,1],[5,3],[2,4]],[[1,2],[3,4],[5,6],[1,4],[6,3],[2,5],[6,1],[3,5],[4,2],[1,5],[3,2],[6,4],[3,1],[5,3],[2,6]],[[1,2],[3,4],[5,6],[7,1],[2,3],[4,5],[6,7],[1,3],[2,4],[3,7],[5,2],[1,6],[7,4],[1,5],[2,6],[5,7],[1,3],[3,6],[2,7],[2,7],[5,3],[4,6]]]
		
		#Write Groupnames to Array Matches		
		self.matches_order = []
		for idx, group in enumerate(self.groups):
			self.matches_order.append([])
		#Create Array with Matches
		groupfinish = 0
		for group_round in range(len(self.groups)):
			for group_game in range(int(((len(self.groups[0])-1)*(len(self.groups[0])-2))/2)):
				print(group_round)
				print(group_game)
				print(self.groups[group_round][self.matches_logic[len(self.groups[0])-3][group_game][0]]+'_'+self.groups[group_round][self.matches_logic[len(self.groups[0])-3][group_game][1]])
				self.matches_order[group_round].append(str(self.matches_logic[len(self.groups[0])-3][group_game][0])+'_'+str(self.matches_logic[len(self.groups[0])-3][group_game][1]))
		print(self.matches_order)

		groupfinish = 0	
		#self.mixedenable = 'true' #Temporaer
		starttime_current = self.starttime
		self.order = []
		#for Mixed Group Gameplay
		if self.mixedenable == 'true':
			print('Mixed Group Gameplay choosed')
			for game in range(1,int(self.game_count)+1):
				startfield = [get_column_letter(self.startfield_group[0]),int(self.startfield_group[1])+groupfinish*(len(self.groups[0])-1)]
				print(startfield)

				#Get the current game from the group, group number is variable groupfinish!!!
				groupgame = int(game-((len(self.groups))*groupfinish))
				#Draw
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)] = game	
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)] = str(starttime_current.hour)+':'+str(starttime_current.minute).zfill(2)
				starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)] = self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[0])]	
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)] = self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[1])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				print('Game %d is %s vs %s' % (groupgame, self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[0])],self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[1])]))
				print('Teams of Group: '+str(groupgame-1)+' numbers : '+str(self.matches_order[groupgame-1][groupfinish].split('_')[0]+' and '+self.matches_order[groupgame-1][groupfinish].split('_')[1]))
				self.order.append(str(groupgame-1)+'-'+str(self.matches_order[groupgame-1][groupfinish].split('_')[0])+'-'+str(self.matches_order[groupgame-1][groupfinish].split('_')[1]))	
				if game % (len(self.groups)) == 0:
					groupfinish += 1
					print('Groupfinish %d' % groupfinish)
					if game != self.game_count:
						self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame)] = '0:'+str(self.breaktime).zfill(2)
						starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime))	
				if game == 1:
					#change column dimensions
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0]))].width = 3
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+1))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+1)].width = 7
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+2))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+2)].width = 20
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+3))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+3)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+4))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+4)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+5))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+5)].width = 20

		#for non Mixed Group Gameplay
		else:
			print('Non Mixed Group Gameplay choosed')
			for game in range(1, int(self.game_count)+1):
				#Startfield
				startfield = [get_column_letter(self.startfield_group[0]),int(int(self.startfield_group[1])+groupfinish*(1+((len(self.groups[0])-1)*(len(self.groups[0])-2))/2))]
				print(startfield)
				#Get the current game from the group, group number is variable groupfinish!!!
				#print(game)
				groupgame = int(game-(((len(self.groups[0])-1)*(len(self.groups[0])-2))/2)*groupfinish)
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)] = game	
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)] = str(starttime_current.hour)+':'+str(starttime_current.minute).zfill(2)
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[0])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				print('Game %d is %s vs %s' % (groupgame, self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[0])],self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]))
				print('Teams of Group: '+str(groupfinish)+' with number '+ str(self.matches_order[groupfinish][groupgame-1].split('_')[0])+' and '+str(self.matches_order[groupfinish][groupgame-1].split('_')[1]))
				self.order.append(str(groupfinish)+'-'+ str(self.matches_order[groupfinish][groupgame-1].split('_')[0])+'-'+str(self.matches_order[groupfinish][groupgame-1].split('_')[1]))
				if game % (((len(self.groups[0])-1)*(len(self.groups[0])-2))/2) == 0:
					groupfinish += 1
					print('Groupfinish %d' % groupfinish)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame)] = '0:'+str(self.breaktime).zfill(2)
					starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime))
				if game == 1:
					#change column dimensions
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0]))].width = 3
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+1))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+1)].width = 7
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+2))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+2)].width = 20
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+3))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+3)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+4))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+4)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+5))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+5)].width = 20
		

		#Crisscross games
		#Save Workbook
		print('Gamesorder are:')
		print(self.order)
		self.workbook.save(filename = self.workbook_path)
	
	#Get current Match Teamnames
	def getTeamNames(self, match_count):
		#print(str(match_count)+'old')
		position = match_count
		#match_count = 4		
		#add the amount of empty line feeds the the variable		
		position = match_count + int(position/len(self.groups))
		#if we are at the last line of the box, we have increment because if the line feed, so we have to decrement 
		#for example match_count = 4 then we add above 1 so it is 5, so we have to decrement by 1 to the correct value
		if match_count % len(self.groups) == 0:
			position -= 1
			#print('decrement')
		#print(match_count)
		#print(self.startfield_group[0])
		#print(get_column_letter(self.startfield_group[0]+2))
		current_teamA = self.trackingsheet[get_column_letter(self.startfield_group[0]+2)+str(int(self.startfield_group[1])+position-1)].value
		current_teamB = self.trackingsheet[get_column_letter(self.startfield_group[0]+5)+str(int(self.startfield_group[1])+position-1)].value
		#print('Teamnames are: %s and %s' % (current_teamA, current_teamB))
		return [current_teamA, current_teamB]

	#Write current Match Score to Excel
	def writeMatchValue(self, match_count, value_teamA, value_teamB):
		print('----------writeMatchValue-------------')
		position = match_count
		#self.match_count = 4		
		## Write to Groupgame List ##
		#add the amount of empty line feeds the the variable		
		position = match_count + int(position/len(self.groups))
		#if we are at the last line of the box, we have increment because if the line feed, so we have to decrement 
		#for example match_count = 4 then we add above 1 so it is 5, so we have to decrement by 1 to the correct value
		if match_count % len(self.groups) == 0:
			position -= 1
			print('decrement')
		self.trackingsheet[get_column_letter(self.startfield_group[0]+3)+str(int(self.startfield_group[1])+position-1)] = value_teamA
		self.trackingsheet[get_column_letter(self.startfield_group[0]+4)+str(int(self.startfield_group[1])+position-1)] = value_teamB
		## Wirte to Groupview List ##
		#get teamnames from self.matches array
		group = int(self.order[match_count].split('-')[0])
		teampositionA = int(self.order[match_count].split('-')[1])
		teampositionB = int(self.order[match_count].split('-')[2])
		self.trackingsheet[str(get_column_letter(self.startfield_block[0]+teampositionB)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA))] = str(value_teamA)+':'+str(value_teamB)
		self.trackingsheet[str(get_column_letter(self.startfield_block[0]+teampositionA)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB))] = str(value_teamB)+':'+str(value_teamA)
		#Log
		print('Now had played Group: '+str(group)+' Positions: '+str(teampositionA)+' vs '+str(teampositionB))
		print('Score: '+str(value_teamA)+' written to '+get_column_letter(self.startfield_group[0]+3)+str(int(self.startfield_group[1])+position-1)+' and Score: '+str(value_teamA)+' written to '+get_column_letter(self.startfield_group[0]+4)+str(int(self.startfield_group[1])+position-1))
		print('Value: %s copyied to %s' % (str(value_teamA)+':'+str(value_teamB),str(get_column_letter(self.startfield_block[0]+teampositionB)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA))))
		print('Value: %s copyied to %s' % (str(value_teamA)+':'+str(value_teamB),str(get_column_letter(self.startfield_block[0]+teampositionA)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB))))
		
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
	
	#Define workbookPath
	def defineWorkbookPath(self, workbook_path):
		self.workbook_path = workbook_path

	#Define get WorkbookPath
	def getWorkbookPath(self):
		return str(self.workbook_path)

	#Define get Workbookname
	def getWorkbookName(self):
		return str(self.workbook_path.split('/')[-1])

if __name__ == '__main__':
	class_pytrackscore = pytrackscore()
