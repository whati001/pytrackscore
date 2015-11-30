#!/usr/share/python3
# -*- coding: utf-8 -*-

#Library to Track Value to Excel and read Info from Excel
from openpyxl import *
from openpyxl.styles import *
from openpyxl.cell import *
import subprocess
import datetime
import itertools

class pytrackscore():
	def __init__(self, *args, **kwargs): 
		print('----------pytrackscore started-------------')
		#init function
		#self.workbook_path = None
		#self.defineWorkbook('/Users/Andi/Desktop/Scoreboard/Turnierplan.xlsx')
		#print(self.getWorkbookPath())
		#print(self.getWorkbookName())
		#self.readWorksheetGroups('Groupsname_initialsetup','A1')
		#								Sheetname	starttime, playtime group, breask group, break group cc ps, playtime cc, breaktime cc
		#self.writeWorksheetInitial('Generated_2GROUPS_2TEAMS','8:30','7','5','10','8','10','NEIN',gamemode='MIXEDGROUP')
		#self.getTeamNames(40)
		#self.getTeamNames(1)
		#self.getTeamNames(2)
		#self.getTeamNames(3)
		#self.getTeamNames(4)
        	#self.writeMatchValue(1,1,3)
        	#self.writeMatchValue(2,3,6,'8:35:00')
        	#self.writeMatchValue(3,1,3)
		#self.writeMatchValue(5,20,5)
		#self.writeMatchValue(40,20,45)
		#self.writeMatchValue(41,20,62)
		#self.finishGroupgames()
		#self.finishCrisscross4Group()
		#self.createPositiongames()
		#self.finishGame()
		#self.getScore("Group")
		
	def readWorksheetGroups(self, *args):
		print('----------readWorksheetGroups-------------')
		for idx, arg in enumerate(args):
			if idx == 0:
				self.initialsheet_name = arg
			elif idx == 1:
				self.startfield = arg
			#print(arg)
		
		#print(self.initialsheet_name)
		#Read Workbook
		self.initialsheet = self.workbook[self.initialsheet_name]
		#Read Groups
		#if we got a startfield from outside or not
		if not hasattr(self, 'startfield'):
			self.startfield = 'A1'
			self.idx = 1
		else:
			#if we from outside something, set counter
			self.idx = int(self.startfield[1])

		print(self.idx)
		self.groups = [[]]
		self.pos01 = 0
		self.repeat = 'true'
		while self.repeat == 'true':
			self.fieldname = self.startfield[0]+str(self.idx)
			self.fieldvalue = str(self.initialsheet[self.fieldname].value)
			print('Value form Field '+self.fieldname+' : '+self.fieldvalue)
			self.idx += 1
			self.groups[self.pos01].append(self.fieldvalue)
			#if there are two Fields empty behind another, stop reading
			if self.fieldvalue == 'None':
				del self.groups[self.pos01][-1]
				if str(self.initialsheet[self.startfield[0]+str(self.idx)].value) == 'None':
					self.repeat = 'false'
				else:
					self.pos01 += 1
					self.groups.append([])

		print('All Groups were saved to self.groups.')
		print(self.groups)

	#def writeWorksheetInitial(self, trackingworksheetname,starttime,playtime_group,breaktime_group,**kwargs):
	def writeWorksheetInitial(self, *args, **kwargs):
		print('----------writeWorksheetInitial-------------')
		self.crisscross_enable = 'false'
		for idx,arg in enumerate(args):
			if idx == 0:
				self.trackingworksheetname = arg
			elif idx == 1:
				self.starttime = datetime.datetime.strptime(arg, "%H:%M")
			elif idx == 2:
				self.playtime_group = arg
			elif idx == 3:
				self.breaktime_group = arg
			elif idx == 4:
				self.breaktime_group_crisscross = arg
			elif idx == 5:
				self.playtime_crisscross = arg
			elif idx == 6:
				self.breaktime_crisscross = arg
			elif idx == 7 and arg == 'JA':
				self.crisscross_enable = 'true'
		
		#Check with Gamemode will used
		if kwargs['gamemode'] == 'MIXEDGROUP':
			self.mixedenable = 'true'
		else:
			self.mixedenable = 'false'
		'''
		self.trackingworksheetname = trackingworksheetname
		self.starttime = datetime.datetime.strptime(starttime, "%H:%M")
		self.playtime_group = playtime_group
		self.breaktime_group = breaktime_group'''
		#print(self.workbook.get_sheet_names())
		#self.trackingsheet = self.workbook.remove_sheet(self.workbook.worksheets[2])
		if self.trackingworksheetname not in self.workbook.get_sheet_names():
			print('New worksheet was created with sheetname: '+str(self.trackingworksheetname))
			self.trackingsheet = self.workbook.create_sheet(title = self.trackingworksheetname)
		else:
			print('Existing worksheet used with sheetname: '+str(self.trackingworksheetname))
			self.trackingsheet = self.workbook[self.trackingworksheetname]
		#self.workbook.save(filename = self.workbook_path)

		#Initial Drawing
		#Groupcolors
		self.groupcolors = ['ebf1de','fdebda','dbeef4','e6e0ec']
		#Draw GroupMatches
		self.matches_logic_group = [[[1,2],[2,1]],[[1,2],[3,1],[2,3]],[[1,2],[3,4],[1,3],[2,4],[4,1],[3,2]],[[1,2],[3,4],[5,1],[2,3],[4,5],[1,3],[2,5],[4,1],[5,3],[2,4]],[[1,2],[3,4],[5,6],[1,4],[6,3],[2,5],[6,1],[3,5],[4,2],[1,5],[3,2],[6,4],[3,1],[5,3],[2,6]],[[1,2],[3,4],[5,6],[7,1],[2,3],[4,5],[6,7],[1,3],[2,4],[3,7],[5,2],[1,6],[7,4],[1,5],[2,6],[5,7],[1,3],[3,6],[2,7],[2,7],[5,3],[4,6]]]
		self.gamescore_fields=[]

		#Draw GroupGames to Excel
		self.startfield_group = [ord('B')-64,'4']
		if int(self.startfield_group[1]) > 3:
			#Draw Title "Gruppenspiel"
			self.trackingsheet[get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-2)] = 'Gruppenspiele'
			self.trackingsheet.merge_cells(get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-2)+':'+get_column_letter(self.startfield_group[0]+5)+str(int(self.startfield_group[1])-2))
			self.trackingsheet[get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-2)].font = Font(bold=True)
			self.trackingsheet[get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-2)].alignment = Alignment(horizontal='center')	
			#Draw groupplaytime and infotext
			self.trackingsheet[get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-1)] = self.playtime_group+' Minuten Spieldauer pro Gruppenspiel mit ' + self.breaktime_group+' Pause'
			self.trackingsheet[get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-1)].alignment = Alignment(horizontal='center')	
			self.trackingsheet.merge_cells(get_column_letter(self.startfield_group[0])+str(int(self.startfield_group[1])-1)+':'+get_column_letter(self.startfield_group[0]+5)+str(int(self.startfield_group[1])-1))

		self.game_count = len(self.groups)*((len(self.groups[0])-1)*(len(self.groups[0])-2))/2
		print('In InitalGroup Sheet were : '+str(len(self.groups))+' Groups with each '+str(len(self.groups[0]))+' Teams')
		print('Total there are : %d GroupGames' % (self.game_count))
		
		#Write Groupnames to Array Matches		
		self.matches_order = []
		for idx, group in enumerate(self.groups):
			self.matches_order.append([])
		#Create Array with Matches
		groupfinish = 0
		for group_round in range(len(self.groups)):
			for group_game in range(int(((len(self.groups[0])-1)*(len(self.groups[0])-2))/2)):
				print('Game %d from Group %d is %s' %(group_game,group_round,str(self.groups[group_round][self.matches_logic_group[len(self.groups[0])-3][group_game][0]]+'_'+self.groups[group_round][self.matches_logic_group[len(self.groups[0])-3][group_game][1]])))
				self.matches_order[group_round].append(str(self.matches_logic_group[len(self.groups[0])-3][group_game][0])+'_'+str(self.matches_logic_group[len(self.groups[0])-3][group_game][1]))
		print(self.matches_order)

		groupfinish = 0	
		#self.mixedenable = 'true' #Temporaer
		starttime_current = self.starttime
		self.order = []
		#for Mixed Group Gameplay
		if self.mixedenable == 'true':
			print('Mixed Group Gameplay choosed')
			for game in range(1,int(self.game_count)+1):
				self.gamescore_fields.append([])
				startfield = [get_column_letter(self.startfield_group[0]),int(self.startfield_group[1])+groupfinish*(len(self.groups)+1)+1]
				print(startfield)

				#Get the current game from the group, group number is variable groupfinish!!!
				groupgame = int(game-((len(self.groups))*groupfinish))
				#Draw
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)] = game	
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)] = str(starttime_current.hour).zfill(2)+':'+str(starttime_current.minute).zfill(2)
				starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_group))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)] = self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[0])]	
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#Save Field location
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1))
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1))
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1))
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1))								
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)] = self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[1])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupgame-1])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				print('Game %d is %s vs %s' % (groupgame, self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[0])],self.groups[groupgame-1][int(self.matches_order[groupgame-1][groupfinish].split('_')[1])]))
				print('Teams of Group: '+str(groupgame-1)+' numbers : '+str(self.matches_order[groupgame-1][groupfinish].split('_')[0]+' and '+self.matches_order[groupgame-1][groupfinish].split('_')[1]))
				self.order.append(str(groupgame-1)+'-'+str(self.matches_order[groupgame-1][groupfinish].split('_')[0])+'-'+str(self.matches_order[groupgame-1][groupfinish].split('_')[1]))	
				if game % (len(self.groups)) == 0:
					groupfinish += 1
					print('Groupfinish %d' % groupfinish)
					if game != self.game_count:
						self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame)] = '0:'+str(self.breaktime_group).zfill(2)+ ' Minuten Pause'
						starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_group))
				if game == 1:
					#change column dimensions
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0]))].width = 3
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+1))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+1)].width = 7
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+2))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+2)].width = 20
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+3))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+3)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+4))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+4)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+5))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+5)].width = 20

			#set startfield for following Games, if they are enabled
			self.startfield_following = [get_column_letter(self.startfield_group[0]),int(self.startfield_group[1])+groupfinish*(len(self.groups)+1)+4]


		#for non Mixed Group Gameplay
		else:
			print('Non Mixed Group Gameplay choosed')
			blankline = 0
			blankline_counter = 0
			for game in range(1, int(self.game_count)+1):
				self.gamescore_fields.append([])
				blankline_counter += 1
				#Startfield
				###startfield = [get_column_letter(self.startfield_group[0]),int(int(self.startfield_group[1])+groupfinish*(1+((len(self.groups[0])-1)*(len(self.groups[0])-2))/2))]
				startfield = [get_column_letter(self.startfield_group[0]),int(int(self.startfield_group[1])+blankline*5)]
				
				print(startfield)
				'''#Get the current game from the group, group number is variable groupfinish!!!
				#print(game)
				
				print(groupgame)
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)] = game	
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[startfield[0]+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)] = str(starttime_current.hour)+':'+str(starttime_current.minute).zfill(2)
				starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_group))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[0])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+groupgame-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				print('Game %d is %s vs %s' % (groupgame, self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[0])],self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]))
				#print('Teams of Group: '+str(groupfinish)+' with number '+ str(self.matches_order[groupfinish][groupgame-1].split('_')[0])+' and '+str(self.matches_order[groupfinish][groupgame-1].split('_')[1]))
				'''
				groupgame = int(game-(((len(self.groups[0])-1)*(len(self.groups[0])-2))/2)*groupfinish)
				self.trackingsheet[startfield[0]+str(startfield[1]+blankline_counter)] = game	
				self.trackingsheet[startfield[0]+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[startfield[0]+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+blankline_counter)] = str(starttime_current.hour).zfill(2)+':'+str(starttime_current.minute).zfill(2)
				starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_group))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+blankline_counter)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[0])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))			
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#Save Field Location
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+2)+str(startfield[1]+blankline_counter))
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+3)+str(startfield[1]+blankline_counter))
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+4)+str(startfield[1]+blankline_counter))	
				self.gamescore_fields[game-1].append(get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+blankline_counter))
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+blankline_counter)] = self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+blankline_counter)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[groupfinish])
				self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+5)+str(startfield[1]+blankline_counter)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				print('Game %d is %s vs %s' % (groupgame, self.groups[groupfinish][int(self.matches_order[groupfinish][blankline_counter].split('_')[0])],self.groups[groupfinish][int(self.matches_order[groupfinish][groupgame-1].split('_')[1])]))
				#print('Teams of Group: '+str(groupfinish)+' with number '+ str(self.matches_order[groupfinish][groupgame-1].split('_')[0])+' and '+str(self.matches_order[groupfinish][groupgame-1].split('_')[1]))
				
				if game % 4 == 0 and game != self.game_count:
					#if game != self.game_count:
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+blankline_counter+1)] = '0:'+str(self.breaktime_group).zfill(2)+ ' Pause'
					starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_group))
					blankline += 1
					blankline_counter = 0
					
				self.order.append(str(groupfinish)+'-'+ str(self.matches_order[groupfinish][groupgame-1].split('_')[0])+'-'+str(self.matches_order[groupfinish][groupgame-1].split('_')[1]))
				if game % (((len(self.groups[0])-1)*(len(self.groups[0])-2))/2) == 0:
					groupfinish += 1
					print('Groupfinish %d' % groupfinish)
					###if game != self.game_count:
					###	self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+1)+str(startfield[1]+groupgame)] = '0:'+str(self.breaktime_group).zfill(2)+ ' Pause'
					###	starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_group))

				if game == 1:
					#change column dimensions
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0]))].width = 3
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+1))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+1)].width = 7
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+2))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+2)].width = 20
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+3))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+3)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+4))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+4)].width = 10
					print('Change Column Dimensions for : '+get_column_letter(column_index_from_string(startfield[0])+5))
					self.trackingsheet.column_dimensions[get_column_letter(column_index_from_string(startfield[0])+5)].width = 20
			#set startfield for following Games, if they are enabled
			###self.startfield_following = [get_column_letter(self.startfield_group[0]),int(int(self.startfield_group[1])+groupfinish*(1+((len(self.groups[0])-1)*(len(self.groups[0])-2))/2))+3]
			self.startfield_following = [get_column_letter(self.startfield_group[0]),int(int(self.startfield_group[1])+blankline+self.game_count+5)]

		
		#All Groups with there teams
		self.startfield_block = [ord('J')-64,'4']	
		for count01, group in enumerate(self.groups):
			#All teams in one Group
			#print(group)
			startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
			for count02, team in enumerate(group):
				for count03 in range(len(self.groups[count01])):
					print("Fill : %s" % get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count03)+str(int(startfield[1])+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#One Team
				#Teams vertical
				print('Write '+str(team)+' to '+str(startfield[0])+str(startfield[1]+count02))
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)] = team
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].font = Font(bold=True)
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
				self.trackingsheet[str(startfield[0])+str(startfield[1]+count02)].border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
				#if the Groupname is written start with horicontal writting
				if count02 != 0:
					#Draw Teamposition after the groupgames to the right of the table, print only 1 - self.groups[X]-1 (first is only Groupname) count
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+8)+str(startfield[1]+count02)] = count02
					print('Write '+str(team)+' to '+chr(ord(startfield[0])+count02)+str(startfield[1]))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])] = team
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].fill = PatternFill(patternType='solid', start_color=self.groupcolors[count01])
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(startfield[1])].border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))

					#Write X to the SAME TEAM fields
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)] = 'X'
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].fill = PatternFill(patternType='solid', start_color='cccccc')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+count02)+str(int(startfield[1])+count02)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				if count01 == 0:
					#First group loop modify column width
					print('Change Column Dimensions to 20 for : '+str(chr(ord(startfield[0])+count02)))
					self.trackingsheet.column_dimensions[chr(ord(startfield[0])+count02)].width = 25
				if count02 == 0:
					#Draw Scores for Teams, left from All teams in Group Table
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1])] = "Siege"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1])].alignment = Alignment(horizontal='center')	
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1])] = "Unentschieden"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1])] = "Niederlage"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1])] = "Tore"
					self.trackingsheet.merge_cells(get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1])+':'+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+4)+str(startfield[1]))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1])] = "Punkte"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1])].alignment = Alignment(horizontal='center')

					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+8)+str(startfield[1])] = "Platzierung"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+8)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+8)+str(startfield[1])].alignment = Alignment(horizontal='center')							
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9)+str(startfield[1])] = "Teamname"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+10)+str(startfield[1])] = "Spiele"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+10)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+10)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+11)+str(startfield[1])] = "Siege"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+11)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+11)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+12)+str(startfield[1])] = "Unentschieden"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+12)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+12)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+13)+str(startfield[1])] = "Niederlage"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+13)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+13)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+14)+str(startfield[1])] = "Tore"
					self.trackingsheet.merge_cells(get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+14)+str(startfield[1])+':'+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+15)+str(startfield[1]))
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+14)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+14)+str(startfield[1])].alignment = Alignment(horizontal='center')
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+16)+str(startfield[1])] = "Punkte"
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+16)+str(startfield[1])].font = Font(bold=True)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+16)+str(startfield[1])].alignment = Alignment(horizontal='center')
	
		print('Gamesorder for Groups are:')
		print(self.order)

		self.block_count = [[[2],[3],[4],[2,3],[3,3],[2,2,3]],[[4,4],[3,3,4],[4,4,4],[4,3,3,4],[4,4,4,4],[4,4,4,2,4]]]

		#Crisscross games
		self.game_count_cc = 0
		if self.crisscross_enable == 'true':
			print('Draw CrissCross Games')
			#Groupcolors
			self.startfield_cc =  self.startfield_following
			#self.groupcolors = ['ebf1de','fdebda','dbeef4','e6e0ec']
			starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_group_crisscross))
			print('Starttime for CC is : '+str(starttime_current))
			#Draw Title "CrissCrossgames"
			self.trackingsheet[self.startfield_cc[0]+str(int(self.startfield_cc[1])-3)] = 'Kreuzspiele'
			self.trackingsheet.merge_cells(self.startfield_cc[0]+str(int(self.startfield_cc[1])-3)+':'+get_column_letter(column_index_from_string(self.startfield_cc[0])+5)+str(int(self.startfield_cc[1])-3))
			self.trackingsheet[self.startfield_cc[0]+str(int(self.startfield_cc[1])-3)].font = Font(bold=True)
			self.trackingsheet[self.startfield_cc[0]+str(int(self.startfield_cc[1])-3)].alignment = Alignment(horizontal='center')	
			#Draw Crisscross and infotext
			self.trackingsheet[self.startfield_cc[0]+str(int(self.startfield_cc[1])-2)] = self.playtime_crisscross+' Minuten Spieldauer pro Kreuzspiel mit ' + self.breaktime_crisscross+' Minuten Pause'
			self.trackingsheet[self.startfield_cc[0]+str(int(self.startfield_cc[1])-2)].alignment = Alignment(horizontal='center')	
			self.trackingsheet.merge_cells(self.startfield_cc[0]+str(int(self.startfield_cc[1])-2)+':'+get_column_letter(column_index_from_string(self.startfield_cc[0])+5)+str(int(self.startfield_cc[1])-2))
			
			#Draw CrissCrossgames
			#CrissCrosscolors
			self.crisscrosscolors = ['dceef4','e6e0ec','8db4e3','c4d69b']
			blankline = 0
			cc_group_block_count = 0
			#print(self.matches_order)
			#print(self.game_count)
			#self.groups[0]-1:ignore Groupname
			self.game_count_cc = (len(self.groups[0])-1)*int(len(self.groups)/2)
			if len(self.groups) > 2 :
				game_count_cc_initial = self.game_count_cc
				self.game_count_cc = self.game_count_cc + 4
			else:
				game_count_cc_initial = self.game_count_cc
			#print(self.game_count_cc)
			if len(self.groups) > 2:
				for cc_group_count in range(1,self.game_count_cc+1):
						self.gamescore_fields.append([])
						#print(cc_group_count)
						cc_group_block_count += 1
						###Make format for Crisscross fields
						if cc_group_count <= game_count_cc_initial:
							if cc_group_count % 2 == 0:
								color = self.crisscrosscolors[1]
							elif cc_group_count % 2 == 1:
								color = self.crisscrosscolors[0]
						else: 
							if (cc_group_count > game_count_cc_initial) and (cc_group_count <= game_count_cc_initial+2):
								color = self.crisscrosscolors[2]
							else:
								color = self.crisscrosscolors[3]

				
						#Draw format with color info from above
						for idx in range(6):
							#Format Crisscross field
							self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline)].fill = PatternFill(patternType='solid', start_color=color)
							self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
							#Save Field Locales
							if idx >= 2 and idx <=5:
								self.gamescore_fields[int(self.game_count + cc_group_count-1)].append(get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline))						
					
						###Value in Fields
						self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0]))+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline)] = self.game_count + cc_group_count
						self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+1)+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline)]  = str(starttime_current.hour).zfill(2)+':'+str(starttime_current.minute).zfill(2)
						starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_crisscross))
						#Check if the block size is reached and, then draw the break info and increment blankline
						if (cc_group_block_count == self.block_count[int(len(self.groups)/2-1)][len(self.groups[0])-3][blankline]):
							blankline += 1
							#print(blankline)
							cc_group_block_count = 0
							if cc_group_count != self.game_count_cc:
								self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+1)+str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline)] = '0:'+str(self.breaktime_crisscross).zfill(2)+ ' Minuten Pause'
								starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_crisscross))
							#Set self.startfield_cc for Positiongames
							else:
								self.startfield_pg = [get_column_letter(column_index_from_string(self.startfield_cc[0])),str(int(self.startfield_cc[1])+(cc_group_count)-1+blankline+4)]
							#print('Create blank line')		
			else:
				for cc_group_count in range(1 ,3):
					self.gamescore_fields.append([])
					#print(cc_group_count)
					###Make format for Crisscross fields
					if cc_group_count % 2 == 0:
						color = self.crisscrosscolors[1]
					elif cc_group_count % 2 == 1:
						color = self.crisscrosscolors[0]
			
					#Draw format with color info from above
					for idx in range(6):
						#Format Crisscross field
						self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1)].fill = PatternFill(patternType='solid', start_color=color)
						self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
						#Save Field Locales
						if idx >= 2 and idx <=5:
							self.gamescore_fields[int(self.game_count + cc_group_count-1)].append(get_column_letter(column_index_from_string(self.startfield_cc[0])+idx)+str(int(self.startfield_cc[1])+(cc_group_count)-1))						
				
					###Value in Fields
					self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0]))+str(int(self.startfield_cc[1])+(cc_group_count)-1)] = self.game_count + cc_group_count
					self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_cc[0])+1)+str(int(self.startfield_cc[1])+(cc_group_count)-1)]  = str(starttime_current.hour).zfill(2)+':'+str(starttime_current.minute).zfill(2)
					starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_crisscross))
					self.startfield_pg = [get_column_letter(column_index_from_string(self.startfield_cc[0])),str(int(self.startfield_cc[1])+(cc_group_count)+4)]
					self.game_count_cc = 2
		
		#Draw Positiongames
		#Positiongamescolor
		#self.crisscrosscolors = ['dceef4','e6e0ec','8db4e3','c4d69b']			
		print('Draw Position Games')
		if self.crisscross_enable == 'true':
			self.startfield_pg = self.startfield_pg
		else:
			self.startfield_pg = self.startfield_following
		print(self.startfield_pg)
		blankline = 0
		pg_group_block_count = 0
		starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_group_crisscross))
		print('Starttime for PG is : '+str(starttime_current)+' and startfield is: ')
		#print(self.startfield_pg)
		#print(self.startfield_cc)
		#Draw Title "Positiongames"
		self.trackingsheet[self.startfield_pg[0]+str(int(self.startfield_pg[1])-3)] = 'Positionsspiele'
		self.trackingsheet.merge_cells(self.startfield_pg[0]+str(int(self.startfield_pg[1])-3)+':'+get_column_letter(column_index_from_string(self.startfield_pg[0])+5)+str(int(self.startfield_pg[1])-3))
		self.trackingsheet[self.startfield_pg[0]+str(int(self.startfield_pg[1])-3)].font = Font(bold=True)
		self.trackingsheet[self.startfield_pg[0]+str(int(self.startfield_pg[1])-3)].alignment = Alignment(horizontal='center')	
		#Draw Positiongames and infotext
		self.trackingsheet[self.startfield_pg[0]+str(int(self.startfield_pg[1])-2)] = self.playtime_crisscross+' Minuten Spieldauer pro Positionsspiel mit ' + self.breaktime_crisscross+' Minuten Pause'
		self.trackingsheet[self.startfield_pg[0]+str(int(self.startfield_pg[1])-2)].alignment = Alignment(horizontal='center')	
		self.trackingsheet.merge_cells(self.startfield_pg[0]+str(int(self.startfield_pg[1])-2)+':'+get_column_letter(column_index_from_string(self.startfield_pg[0])+5)+str(int(self.startfield_pg[1])-2))


		###Value in Fields
		for pg_count in range(int(len(self.groups)*(len(self.groups[0])-1)/2)):
			self.gamescore_fields.append([])
			pg_group_block_count += 1
			for idx in range(6):
				#Format Positiongames field
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+idx)+str(int(self.startfield_pg[1])+(pg_count)+blankline)].fill = PatternFill(patternType='solid', start_color='FAFAFA')
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+idx)+str(int(self.startfield_pg[1])+(pg_count)+blankline)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
				#Save Field Locales
				if idx >= 2 and idx <=5:
					self.gamescore_fields[int(self.game_count + self.game_count_cc + pg_count)].append(get_column_letter(column_index_from_string(self.startfield_pg[0])+idx)+str(int(self.startfield_pg[1])+(pg_count)+blankline))			
								
			###Value in Fields
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_pg[1])+(pg_count)+blankline)] = self.game_count + self.game_count_cc + pg_count + 1
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+1)+str(int(self.startfield_pg[1])+(pg_count)+blankline)]  = str(starttime_current.hour).zfill(2)+':'+str(starttime_current.minute).zfill(2)
			starttime_current = starttime_current + datetime.timedelta(minutes = int(self.playtime_crisscross))
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+6)+str(int(self.startfield_pg[1])+(pg_count)+blankline)] = 'Plätze '+str(len(self.groups)*(len(self.groups[0])-1)-pg_count*2)+' und '+str(len(self.groups)*(len(self.groups[0])-1)-(pg_count+1)*2+1)
			#Check if the block size is reached and, then draw the break info and increment blankline
			if (pg_group_block_count == self.block_count[int(len(self.groups)/2-1)][len(self.groups[0])-3][blankline]):
				blankline += 1
				#print(blankline)
				pg_group_block_count = 0
				if pg_count != int(len(self.groups)*(len(self.groups[0])-1)/2-1):
					self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+1)+str(int(self.startfield_pg[1])+(pg_count)+blankline)] = '0:'+str(self.breaktime_crisscross).zfill(2)+ ' Minuten Pause'
					starttime_current = starttime_current + datetime.timedelta(minutes = int(self.breaktime_crisscross))
				#Set self.startfield_pg for Positiongames
				else:
					self.startfield_finally = [get_column_letter(column_index_from_string(self.startfield_pg[0])),str(int(self.startfield_pg[1])+(pg_count)+blankline+4)]
					self.pg_count = pg_count
				#print('Create blank line')	
		self.game_count_pg = pg_count+1
						
		#Draw Teampositions
		#Teampsoitioncolors
		self.teampositioncolor = ['F5A9A9','A9E2F3']			
		print('Draw Finally TeamPosition')
		#Draw Title "TeamPosition"
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_finally[1])-3)] = 'Teampositionen'
		self.trackingsheet.merge_cells(get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_finally[1])-3)+':'+get_column_letter(column_index_from_string(self.startfield_finally[0])+5)+str(int(self.startfield_finally[1])-3))
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_finally[1])-3)].font = Font(bold=True)
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_finally[1])-3)].alignment = Alignment(horizontal='center')	
		
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0]))+str(int(self.startfield_finally[1])-1)] = 'Position'
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+1)+str(int(self.startfield_finally[1])-1)] = 'Teamname'
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+2)+str(int(self.startfield_finally[1])-1)] = 'Geschossene Tore'
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+3)+str(int(self.startfield_finally[1])-1)] = 'Bekommene Tore'
		self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_pg[0])+4)+str(int(self.startfield_finally[1])-1)] = 'Tor Differenze'

		###Value in Fields
		for ft_count in range(1,int(len(self.groups)*(len(self.groups[0])-1))+1):
			if ft_count %2 == 0:
				color = self.teampositioncolor[0]
			else:
				color = self.teampositioncolor[1]
				
			for idx in range(5):
				#Format Positiongames field
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+idx)+str(int(self.startfield_finally[1])+(ft_count)-1)].fill = PatternFill(patternType='solid', start_color=color)
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+idx)+str(int(self.startfield_finally[1])+(ft_count)-1)].border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
								
			###Value in Fields
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0]))+str(int(self.startfield_finally[1])+(ft_count)-1)] = ft_count
					
		print(self.gamescore_fields)	
			
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
	
	#Get current Match Teamnames
	def getTeamNames(self, match_count):
		print('----------getTeamNames-------------')
		#print(str(match_count)+'old')
		print('Gamecount: Group: '+str(self.game_count)+' CC: '+ str(self.game_count_cc)+' Platzgame: '+ str(self.game_count_pg))		
		position = match_count
		#match_count = 4		
		#add the amount of empty line feeds the the variable		
		position = match_count + int(position/len(self.groups))
		#if we are at the last line of the box, we have increment because if the line feed, so we have to decrement 
		#for example match_count = 4 then we add above 1 so it is 5, so we have to decrement by 1 to the correct value
		if match_count % len(self.groups) == 0:
			position -= 1
			#print('decrement')
		#print(match_count)
		#print(self.startfield_group[0])
		#print(get_column_letter(self.startfield_group[0]+2))
		if match_count <= self.game_count + self.game_count_cc + self.game_count_pg:
			current_teamA = self.trackingsheet[self.gamescore_fields[match_count-1][0]].value
			current_teamB = self.trackingsheet[self.gamescore_fields[match_count-1][3]].value
			if current_teamA == None and current_teamB == None:
				if match_count > self.game_count:
					current_teamA = 'Kreuzspiele folgen'
					current_teamB = 'Kreuzspiele folgen'
				if match_count > self.game_count + self.game_count_cc:
					current_teamA = 'Positionsspiele folgen'
					current_teamB = 'Positionsspiele folgen'
		else:
			current_teamA = 'Turnier fertig'
			current_teamB = 'Turnier fertig'

			
		print('Teamnames for game %s are: %s and %s' % (str(match_count),current_teamA, current_teamB))
		return [current_teamA, current_teamB]

	#Write current Match Score to Excel
	def writeMatchValue(self, *args):
		print('----------writeMatchValue-------------')
		if len(args) < 3:
			raise Exception('Function need minimal three arguments [match_count, value_teamA, value_teamB, (gametelta)]')
		for idx,arg in enumerate(args):
			if idx == 0:
				position = arg
				match_count = arg
				
			##Write to Groupgame List##
			elif idx == 1:
				self.trackingsheet[self.gamescore_fields[match_count-1][1]] = arg
				value_teamA = arg
			elif idx == 2:		
				self.trackingsheet[self.gamescore_fields[match_count-1][2]] = arg
				value_teamB = arg
			elif idx == 3:
				self.gamedelta = arg
		
		#Log
		print('Now had played game : '+ str(match_count) +' with '+str(value_teamA)+' to '+str(value_teamB))
		
		#get_column_letter(column_index_from_string
		#If gamedelta is enabled, do it
		if hasattr(self, 'gamedelta'):
			print('Gamedelta enabled for matchcount %d' %match_count)
			#print(self.gamescore_fields[match_count-1][1])
			#Get realtime from scoreboard
			realtime = int(self.gamedelta.split(':')[0])*60 + int(self.gamedelta.split(':')[1])
			#Get realtime from system
			#realtime = str(datetime.datetime.now().hour*60+datetime.datetime.now().minute)
			#print(self.gamescore_fields)
			#print(len(self.gamescore_fields))
			#calculate timedelta
			generatetime_string = self.trackingsheet[get_column_letter(column_index_from_string(str(self.gamescore_fields[match_count-1][1])[0])-2)+str(self.gamescore_fields[match_count-1][1])[1:]].value
			print(generatetime_string)
			generatetime = int(generatetime_string.split(':')[0])*60+int(generatetime_string.split(':')[1])
			#print('generatetime ' + str(generatetime))
			#print('realtime     ' + str(realtime))
			self.gamedelta = realtime - generatetime
			#print(self.gamedelta)		
			#print('Timedelte are Minutes: ' + str(self.gamedelta))

			for idx in range(match_count, len(self.gamescore_fields)+1):
				#print(str(idx) +' is '+ str(get_column_letter(column_index_from_string(str(self.gamescore_fields[idx-1][1])[0])-2)+str(self.gamescore_fields[idx-1][1])[1:]))
				#print(get_column_letter(column_index_from_string(str(self.gamescore_fields[match_count-1][1])[0])-1))
				generatetime_string = self.trackingsheet[get_column_letter(column_index_from_string(str(self.gamescore_fields[idx-1][1])[0])-2)+str(self.gamescore_fields[idx-1][1])[1:]].value
				#print(generatetime_string)
				#print(self.gamedelta)
				generatetime = int(generatetime_string.split(':')[0])*60+int(generatetime_string.split(':')[1])
				#print('generatetime ' + str(generatetime))
				print('New starttime for game '+str(idx) + ' is '+str(datetime.timedelta(seconds=(generatetime+self.gamedelta)))[2:])
				self.trackingsheet[str(get_column_letter(column_index_from_string(str(self.gamescore_fields[idx-1][1])[0])-2)+str(self.gamescore_fields[idx-1][1])[1:])] = str(datetime.timedelta(seconds=(generatetime+self.gamedelta)))[2:]
				
			#Delete Attribute, so we can enable the delta function for seprate games
			delattr(self, 'gamedelta')
		
		
		## Wirte to Groupview List ##
		#get teamnames from self.matches array#
		if match_count <= self.game_count:
			group = int(self.order[match_count-1].split('-')[0])
			teampositionA = int(self.order[match_count-1].split('-')[1])
			teampositionB = int(self.order[match_count-1].split('-')[2])
		
			#LEFT TEAM, Groupgames LIST
			self.trackingsheet[str(get_column_letter(self.startfield_block[0]+teampositionB)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA))] = str(value_teamA)+':'+str(value_teamB)
			#count goals LEFT TEAM
			var_current_goal_make = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value	
			if var_current_goal_make is None:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = value_teamA
			else:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_goal_make) + value_teamA
		
			var_current_goal_get = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value			
			if var_current_goal_get is None:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = value_teamB
			else:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_goal_get) + value_teamB

			#WIN
			if value_teamA > value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 3
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_score_total) + 3
			#UNDECIDE
			elif value_teamA == value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_score_total) + 1

			#LOSS
			elif value_teamA < value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 0
			
			#self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA)] = 'TEST'
		
			#RIGHT TEAM, Groupgames LIST
			self.trackingsheet[str(get_column_letter(self.startfield_block[0]+teampositionA)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB))] = str(value_teamB)+':'+str(value_teamA)
			#count goals LEFT TEAM
			var_current_goal_make = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value	
			if var_current_goal_make is None:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = value_teamB
			else:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+3)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_goal_make) + value_teamB
		
			var_current_goal_get = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value			
			if var_current_goal_get is None:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = value_teamA
			else:
				self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+4)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_goal_get) + value_teamA

			#LOSS
			if value_teamA > value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 0

			#UNDECIDE
			elif value_teamA == value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+1)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_score_total) + 1

			#WIN
			elif value_teamA < value_teamB:
				var_current_score = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+2)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				var_current_score_total = self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)].value
				if var_current_score is None:
					#print('Is None') no value in Cell
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 1
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0]))+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_score) + 1
				if var_current_score_total is None:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = 3
				else:
					self.trackingsheet[get_column_letter(self.startfield_block[0]+len(self.groups[0])+5)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB)] = int(var_current_score_total) + 3
		
			#Log
			print('Value: %s copyied to %s' % (str(value_teamA)+':'+str(value_teamB),str(get_column_letter(self.startfield_block[0]+teampositionB)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionA))))
			print('Value: %s copyied to %s' % (str(value_teamA)+':'+str(value_teamB),str(get_column_letter(self.startfield_block[0]+teampositionA)+str(int(self.startfield_block[1])+int(group)*(len(self.groups[int(group)])+1)+teampositionB))))
		
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
		
	#Define complete Groupgames
	def finishGroupgames(self):
		print('----------FinishGroupgames-------------')
		self.startfield_block = [ord('J')-64,'4']	
		#self.groups = [['Gruppe A', 'Black Swanz', 'JVP Hofstätten', 'Kugelrund', 'Die Crew', 'SHT'], ['Gruppe B', 'Fireball Hafning', 'Jerich Dispo', 'Bum - Bum', 'Jerich Hasen', 'Blue Eyes Racing'], ['Gruppe C', 'der Böse Wolf und seine Geißlein', 'BELLAFFAIR.AT', 'HFC Ballerinas', "GH Baumi's Bubble Trouble", 'Ball Buzztards'], ['Gruppe D', 'Schlümpfe', 'FC Ambrosi', 'S-TEC Fighters', 'Jerich Hoolts Durch', 'Formation 88']]
		self.group_scores = []
		#self.trackingsheet = self.workbook['Generated_21092015_2227']
		print('All Groupgames played, start finishing Groupgames')
		
		#Read the Info from Excel Sheet
		for count01 in range(0, len(self.groups)):
			#For every group
			#set startfield of groupblock
			startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
			self.group_scores.append([])
			#self.group_scores[count01].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0]))+str(startfield[1])].value)	#GroupName
			print('startfield: '+str(startfield[0])+str(startfield[1]))
			
			#Start from 1 to len(...) to skip header line
			for count02 in range(1,len(self.groups[0])):
				#For every team in group
				self.group_scores[count01].append([])
				print('Read from Field: '+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1]+count02))
				#decrement by 1 to start from '0'
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0]))+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0]))+str(startfield[1]+count02)].value)				#TeamName
				else:
					self.group_scores[count01][count02-1].append(0)
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1]+count02)].value)	#win
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1]+count02)] = 0
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1]+count02)].value)	#undecide
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+1)+str(startfield[1]+count02)] = 0
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1]+count02)].value)	#los
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+2)+str(startfield[1]+count02)] = 0
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1]+count02)].value)	#make goals
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+3)+str(startfield[1]+count02)] = 0
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+4)+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+4)+str(startfield[1]+count02)].value)	#get goals
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+4)+str(startfield[1]+count02)] = 0
				if self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1]+count02)].value != None:
					self.group_scores[count01][count02-1].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1]+count02)].value)	#total score
				else:
					self.group_scores[count01][count02-1].append(0)
					self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+5)+str(startfield[1]+count02)] = 0
				#To Include the goal difference, make goals minus get goals diffident by 100;;100 is my basis
				self.group_scores[count01][count02-1].append(int(self.group_scores[count01][count02-1][6])+((int(self.group_scores[count01][count02-1][4])-int(self.group_scores[count01][count02-1][5]))/100))	#total score + goals
			
		print(self.group_scores)

		#Sort and wirte Team
		for count01 in range(0, len(self.groups)):
			startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
			current_group = sorted(self.group_scores[count01], key=lambda x : x[7], reverse=True)
			self.group_scores[count01] = current_group
			print(current_group)
			#Start from 1 to len(...) to skip header line
			for count02 in range(0,len(current_group)): 
				#self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+8)+str(startfield[1]+1+count02)] = count02+1
				print(current_group[count02])
				for count03 in range(0,len(current_group[count02])): 
					if count03 == 0:
						print('Write Value: '+str(current_group[count02][count03])+' to Field: '+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02))
						self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02)] = current_group[count02][count03]
					elif count03 == 1:
						print('Write Value: '+str(current_group[count02][1]+current_group[count02][2]+current_group[count02][3])+' to Field: '+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02))
						self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02)] = current_group[count02][1]+current_group[count02][2]+current_group[count02][3]
					else:
						print('Write Value: '+str(current_group[count02][count03-1])+' to Field: '+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02))
						self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9+count03)+str(startfield[1]+1+count02)] = current_group[count02][count03-1]
	


		#print('Gamecount: Group: '+str(self.game_count )+' CC: '+ str(self.game_count_cc )+' Platzgame: '+ str(self.pg_count))
		#Fill first CrissCross table (FIRST CrissCross are the all expect the cc from the cc, the four games between winner and losser from the crisscross
		self.logic_cc = [[['00','11'],['01','10'],['02','12'],['03','13'],['04','14'],['05','15'],['06','16']],[['00','11'],['01','10'],['20','31'],['21','30'],['02','12'],['22','32'],['02','13'],['23','33'],['04','14'],['24','34'],['05','15'],['25','35'],['06','16'],['26','36']]]
		if self.crisscross_enable == 'true':
			#for all first crisscross games
			for count01 in range(int(len(self.groups)/2)*(len(self.groups[0])-1)):
				#Check if 2 or 4 groups, because of selection from the logic_cc
				if int(len(self.groups)) > 2:	
					print('4 Groups')
					print('CC Game '+str(count01)+ ' is ' + str(self.group_scores[int(self.logic_cc[1][count01][0][0])][int(self.logic_cc[1][count01][0][1])][0]) + ' vs '+str(self.group_scores[int(self.logic_cc[1][count01][1][0])][int(self.logic_cc[1][count01][1][1])][0]))
					#print(self.logic_cc[1][count01][0][1])
					#print(self.group_scores[1][count01][0][0])][int(self.logic_cc[1][len(self.groups[0])-2][count01][0][1])][0])
					#print(self.group_scores[1][count01][1][0])][int(self.logic_cc[1][len(self.groups[0])-2][count01][1][1])][0])
					#print(self.group_scores[int(self.logic_cc[1][count01][0][0])][int(self.logic_cc[1][count01][0][1])][0])
					self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][0]] = self.group_scores[int(self.logic_cc[1][count01][0][0])][int(self.logic_cc[1][count01][0][1])][0]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][3]] = self.group_scores[int(self.logic_cc[1][count01][1][0])][int(self.logic_cc[1][count01][1][1])][0]
				else:
					print('2 Groups')
					print('CC Game '+str(count01)+ ' is ' + str(self.group_scores[int(self.logic_cc[0][count01][0][0])][int(self.logic_cc[0][count01][0][1])][0]) + ' vs '+str( self.group_scores[int(self.logic_cc[0][count01][1][0])][int(self.logic_cc[0][count01][1][1])][0]))
					self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][0]] = self.group_scores[int(self.logic_cc[0][count01][0][0])][int(self.logic_cc[0][count01][0][1])][0]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][3]] = self.group_scores[int(self.logic_cc[0][count01][1][0])][int(self.logic_cc[0][count01][1][1])][0]
					if count01 == 1:
						break

		
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
		
	#Define finishCrissCross4Groups
	def finishCrisscross4Group(self):
		print('----------CrissCross4Groups-------------')
		#Fill the last 4 CrissCross Games
		#Read the gamescores from Sheet
		self.crisscross_scores_4group = []
		for count01 in range(4):
			self.crisscross_scores_4group.append([])
			print('Read from CC Game %d' % count01)
			print(self.gamescore_fields[int(self.game_count+self.game_count_cc-4+count01)])
			self.crisscross_scores_4group[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][0]].value)
			self.crisscross_scores_4group[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][1]].value)
			self.crisscross_scores_4group[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][2]].value)
			self.crisscross_scores_4group[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][3]].value)

		#Set the teamnames to the rigth field in Sheet
		if self.crisscross_scores_4group[0][1] > self.crisscross_scores_4group[0][2]:
			#Winner draw to winner
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4)][0]] = self.crisscross_scores_4group[0][0]		
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-2)][3]] = self.crisscross_scores_4group[0][3]
		else:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4)][0]] = self.crisscross_scores_4group[0][3]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-2)][3]] = self.crisscross_scores_4group[0][0]
			
		if self.crisscross_scores_4group[1][1] > self.crisscross_scores_4group[1][2]:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-3)][0]] = self.crisscross_scores_4group[1][0]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-1)][3]] = self.crisscross_scores_4group[1][3]
		else:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-3)][0]] = self.crisscross_scores_4group[1][3]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-1)][3]] = self.crisscross_scores_4group[1][0]	

		if self.crisscross_scores_4group[2][1] > self.crisscross_scores_4group[2][2]:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4)][3]] = self.crisscross_scores_4group[2][0]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-2)][0]] = self.crisscross_scores_4group[2][3]
		else:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4)][3]] = self.crisscross_scores_4group[2][3]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-2)][0]] = self.crisscross_scores_4group[2][0]

		if self.crisscross_scores_4group[3][1] > self.crisscross_scores_4group[3][2]:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-3)][3]] = self.crisscross_scores_4group[3][0]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-1)][0]] = self.crisscross_scores_4group[3][3]
		else:
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-3)][3]] = self.crisscross_scores_4group[3][3]
			self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-1)][0]] = self.crisscross_scores_4group[3][0]
								
		print(self.crisscross_scores_4group)
		
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
		
	#Define createPositiongames
	def createPositiongames(self):
		print('----------createPositiongames-------------')
		#Fill the PositionGames
		self.crisscross_scores = []
		#For four Teams
		#Read value from Excel
		if int(len(self.groups)) > 2:
			self.crisscross_scores = []
			for count01 in range(self.game_count_cc-4):
				self.crisscross_scores.append([])
				print('Read from CC Game %d' % count01)
				#Read the double crisscross games
				if count01 < 4:
					print(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-count01)][0]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4+count01-1)][0]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4+count01-1)][1]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4+count01-1)][2]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc-4+count01-1)][3]].value)
				else:
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][0]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][1]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][2]].value)
					self.crisscross_scores[count01].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][3]].value)
		
			count01 = 1
			while count01 < len(self.crisscross_scores):
				print(str(count01)+str(self.crisscross_scores[-count01]))
				print(str(count01)+str(self.crisscross_scores[-count01-1]))
				print(self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)])
			
				if self.crisscross_scores[-count01][1] > self.crisscross_scores[-count01][2]:
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)][0]] = self.crisscross_scores[-count01][0]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][3]] = self.crisscross_scores[-count01][3]
				else:
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)][0]] = self.crisscross_scores[-count01][3]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][3]] = self.crisscross_scores[-count01][0]
						
				if self.crisscross_scores[-count01-1][1] > self.crisscross_scores[-count01-1][2]:
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)][3]] = self.crisscross_scores[-count01-1][0]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0]] = self.crisscross_scores[-count01-1][3]
				else:
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)][3]] = self.crisscross_scores[-count01-1][3]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0]] = self.crisscross_scores[-count01-1][0]
				
				count01 +=2
		
		else:
			if self.crisscross_enable == 'true':
				print('Positiongames 2 Groups and CC')

				#Read the Info from Excel Sheet
				for count01 in range(0, len(self.groups)):
					#For every group
					#set startfield of groupblock
					startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
					self.crisscross_scores.append([])
					print(self.crisscross_scores)
					print('startfield: '+str(startfield[0])+str(startfield[1]))
					#Start from 1 to len(...) to skip header line
					for count02 in range(1,len(self.groups[0])):
						#For every team in group
						if count02 > 2:
							self.crisscross_scores[count01].append([])
							self.crisscross_scores[count01][count02-3].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9)+str(startfield[1]+count02)].value)	#total score
						#elif count02 == 1:
						#	self.crisscross_scores[count01].append([])
						#	self.crisscross_scores[count01][count02-1].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][0]].value)
						#	self.crisscross_scores[count01][count02-1].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][1]].value)
						#	self.crisscross_scores[count01][count02-1].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][2]].value)
						#	self.crisscross_scores[count01][count02-1].append(self.trackingsheet[self.gamescore_fields[int(self.game_count+count01)][3]].value)
						
				print(self.crisscross_scores)
				
				for count01 in range(1,len(self.crisscross_scores[0])+1):
					#print(count01)
					#print(self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0])
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0]] = self.crisscross_scores[0][-count01][0]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][3]] = self.crisscross_scores[1][-count01][0]
				startfield = int(self.game_count+self.game_count_cc+count01-1)+1
				#print(startfield)
					
				#print('####')
				#print(self.gamescore_fields[int(self.game_count)][0])
				
				if self.trackingsheet[self.gamescore_fields[int(self.game_count)][1]].value > self.trackingsheet[self.gamescore_fields[int(self.game_count)][2]].value:
					self.trackingsheet[self.gamescore_fields[startfield+1][0]] = self.trackingsheet[self.gamescore_fields[int(self.game_count)][0]].value
					self.trackingsheet[self.gamescore_fields[startfield][0]] = self.trackingsheet[self.gamescore_fields[int(self.game_count)][3]].value
				else:
					self.trackingsheet[self.gamescore_fields[startfield+1][0]] = self.trackingsheet[self.gamescore_fields[int(self.game_count)][3]].value
					self.trackingsheet[self.gamescore_fields[startfield][0]] = self.trackingsheet[self.gamescore_fields[int(self.game_count)][0]].value
					
				print(self.gamescore_fields[int(self.game_count+1)])
				if self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][1]].value > self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][2]].value:
					self.trackingsheet[self.gamescore_fields[startfield+1][3]] = self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][0]].value
					self.trackingsheet[self.gamescore_fields[startfield][3]] = self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][3]].value
				else:
					self.trackingsheet[self.gamescore_fields[startfield+1][3]] = self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][3]].value
					self.trackingsheet[self.gamescore_fields[startfield][3]] = self.trackingsheet[self.gamescore_fields[int(self.game_count+1)][0]].value

			else:
				print('Positiongames 2 Groups no CC')
				
				self.group_scores = []
				#Read the Info from Excel Sheet
				for count01 in range(0, len(self.groups)):
					#For every group
					#set startfield of groupblock
					startfield = [get_column_letter(self.startfield_block[0]),int(self.startfield_block[1])+count01*(len(self.groups[count01])+1)]
					self.group_scores.append([])
					#self.group_scores[count01].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0]))+str(startfield[1])].value)	#GroupName
					print('startfield: '+str(startfield[0])+str(startfield[1]))
			
					#Start from 1 to len(...) to skip header line
					for count02 in range(1,len(self.groups[0])):
						#For every team in group
						print('Read from Field: '+get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0]))+str(startfield[1]+count02))
						#decrement by 1 to start from '0'
						self.group_scores[count01].append(self.trackingsheet[get_column_letter(column_index_from_string(startfield[0])+len(self.groups[0])+9)+str(startfield[1]+count02)].value)	#total score
				print(self.group_scores)

				for count01 in range(1,len(self.group_scores[0])+1):
					print(count01)
					#print(self.gamescore_fields[int(self.game_count+self.game_count_cc+count01)])
					print('write : '+str(self.group_scores[0][-count01]) + ' to '+ str(self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0]))
					print('write : '+str(self.group_scores[1][-count01]) + ' to '+ str(self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][3]))
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][0]] = self.group_scores[0][-count01]
					self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+count01-1)][3]] = self.group_scores[1][-count01]
					
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
		
		
		
	#Define finishGame
	def finishGame(self):
		print('----------finishGame-------------')
		#Read and write to sheet
		print(self.startfield_finally)
		places = int((len(self.groups[0])-1)*len(self.groups))
		print(places)
		place_count = 0
		for count01 in range(0,int(places/2)):
			print(count01)
			#print(self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][0])
			#print(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][1]].value)
			#print(self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][2]].value)
			#print(get_column_letter(column_index_from_string(self.startfield_finally[0])+1))
			#print(str(int(self.startfield_finally[1])+count01))
			goals = [0,0,0,0]
			if self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][1]].value > self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][2]].value:
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+1)+str(int(self.startfield_finally[1])+place_count)] = self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][0]].value
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+1)+str(int(self.startfield_finally[1])+place_count+1)] = self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][3]].value
									
			else:
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+1)+str(int(self.startfield_finally[1])+place_count)] = self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][3]].value
				self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+1)+str(int(self.startfield_finally[1])+place_count+1)] = self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][0]].value
						
			#Get the total goal scores	
			for count02 in range(0,len(self.gamescore_fields)):
				#print(self.gamescore_fields[count02][0])
				if self.trackingsheet[self.gamescore_fields[count02][0]].value == self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][3]].value:
					goals[0] = goals[0]+int(self.trackingsheet[self.gamescore_fields[count02][1]].value)
					goals[1] = goals[1]+int(self.trackingsheet[self.gamescore_fields[count02][2]].value)
				elif self.trackingsheet[self.gamescore_fields[count02][3]].value == self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][3]].value:
					goals[0] = goals[0]+int(self.trackingsheet[self.gamescore_fields[count02][2]].value)
					goals[1] = goals[1]+int(self.trackingsheet[self.gamescore_fields[count02][1]].value)
				
				if self.trackingsheet[self.gamescore_fields[count02][0]].value == self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][0]].value:
					goals[2] = goals[2]+int(self.trackingsheet[self.gamescore_fields[count02][1]].value)
					goals[3] = goals[3]+int(self.trackingsheet[self.gamescore_fields[count02][2]].value)
				elif self.trackingsheet[self.gamescore_fields[count02][3]].value == self.trackingsheet[self.gamescore_fields[int(self.game_count+self.game_count_cc+int(places/2)-count01-1)][0]].value:
					goals[2] = goals[2]+int(self.trackingsheet[self.gamescore_fields[count02][2]].value)	
					goals[3] = goals[3]+int(self.trackingsheet[self.gamescore_fields[count02][1]].value)
			
			#Goals First Team			
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+2)+str(int(self.startfield_finally[1])+place_count)] = goals[0]
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+3)+str(int(self.startfield_finally[1])+place_count)] = goals[1]
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+4)+str(int(self.startfield_finally[1])+place_count)] = goals[0] - goals[1]
			
			#Goals Second Team
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+2)+str(int(self.startfield_finally[1])+place_count+1)] = goals[2]
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+3)+str(int(self.startfield_finally[1])+place_count+1)] = goals[3]
			self.trackingsheet[get_column_letter(column_index_from_string(self.startfield_finally[0])+4)+str(int(self.startfield_finally[1])+place_count+1)] = goals[2] - goals[3]
			
			#Increment 2 because we draw two places in every loop
			place_count = place_count + 2
			
			
		#Save Workbook
		self.workbook.save(filename = self.workbook_path)
	#Define set workbook
	def defineWorkbook(self, workbook_path):
		print('----------defineWorkbook-------------')
		self.workbook_path = workbook_path
		#Read Workbook
		if '.xlsx' in self.workbook_path:
			self.workbook = load_workbook(filename = self.workbook_path) 
		else:
			var_convert_command = 'soffice --headless --convert-to xlsx '+self.workbook_path+' --outdir '+self.workbook_path[:self.workbook_path.rfind('/')+1]
			subprocess.Popen(var_convert_command, universal_newlines=True, shell=True)
			print('Inputfile %s converted from xls to xlsx and stored to %s.' %(self.workbook_path,self.workbook_path[:self.workbook_path.rfind('/')+1]))
			self.workbook = load_workbook(filename = self.workbook_path)
		print("Workbook defined")

	#Define crisscross Games
	def crisscrossGamesInitial(self):
		print('----------defineCrissCrossGamesInitial-------------')
		
	#Define get WorkbookPath
	def getWorkbookPath(self):
		print('----------getWorkbookPath-------------')
		return str(self.workbook_path)

	#Define get Workbookname
	def getWorkbookName(self):
		print('----------getWorkbookName-------------')
		return str(self.workbook_path.split('/')[-1])
		
		
	#Define get_score
	def getScore(self, playedmode, mode='array'):
		print('----------getScore-------------')
		#print(mode)
		#Testing set playedmode
		#playedmode = "PG"
		
		self.getscore = None
		
		#Check if Group CC or PG was played for initial load
		if playedmode == "Group":
			var_range = int(["".join(x) for _, x in itertools.groupby(self.gamescore_fields[int(self.game_count-1)][0], key=str.isdigit)][1])-1
			print(var_range)
		elif playedmode == "CC":
			var_range = int(["".join(x) for _, x in itertools.groupby(self.gamescore_fields[int(self.game_count+self.game_count_cc-1)][0], key=str.isdigit)][1])-1
		elif playedmode == "PG":
			var_range = int(["".join(x) for _, x in itertools.groupby(self.gamescore_fields[int(self.game_count+self.game_count_cc+self.game_count_pg-1)][0], key=str.isdigit)][1])-1
		
		#Save values from Startfield to the endfield.
		if mode == 'array':
			self.getscore = [[],[],[]]
			#Save Gameinfos
			self.getscore[0].append(playedmode)
			self.getscore[0].append('MIXEDGROUP' if self.mixedenable == 'true' else 'NON-MIXEDGROUP')
			self.getscore[0].append(str(self.groupcolors).replace("[",'').replace("'","").replace("]","").replace(' ',''))
			#Save Games
			for idx in range(var_range):
				self.getscore[1].append([])
				for idx2 in range(6):
					self.getscore[1][idx].append(self.trackingsheet[get_column_letter(self.startfield_group[0]+idx2)+str(int(self.startfield_group[1])-2+idx)].value)
			
			#Save Groupgameview
			for idx in range(int((len(self.groups[0])+1) * len(self.groups) -1)):
				self.getscore[2].append([])
				for idx2 in range(len(self.groups[0])):
						self.getscore[2][idx].append(self.trackingsheet[get_column_letter(self.startfield_block[0]+idx2)+str(int(self.startfield_block[1])+idx)].value)

		else:
			self.getscore = ""
			#Save Gameinfo
			self.getscore = self.getscore + str(playedmode) +','+ ('MIXEDGROUP' if self.mixedenable == 'true' else 'NON-MIXEDGROUP')+','+ str(self.groupcolors).replace("[",'').replace("'","").replace("]","").replace(' ','')+';'
			#Save Games
			for idx in range(var_range):
				for idx2 in range(6):
					if idx2 != 5:
						self.getscore = self.getscore +str(self.trackingsheet[get_column_letter(self.startfield_group[0]+idx2)+str(int(self.startfield_group[1])-2+idx)].value)+','
					else:
						self.getscore = self.getscore +str(self.trackingsheet[get_column_letter(self.startfield_group[0]+idx2)+str(int(self.startfield_group[1])-2+idx)].value)
				if idx != var_range-1:
					self.getscore = self.getscore + '|'
			self.getscore = self.getscore + ';'
			#Save Groupgamesview
			for idx in range(int((len(self.groups[0])+1) * len(self.groups) -1)):
				for idx2 in range(len(self.groups[0])):
					if idx2 != len(self.groups[0])-1:
						self.getscore = self.getscore +str(self.trackingsheet[get_column_letter(self.startfield_block[0]+idx2)+str(int(self.startfield_block[1])+idx)].value)+','
					else:
						self.getscore = self.getscore +str(self.trackingsheet[get_column_letter(self.startfield_block[0]+idx2)+str(int(self.startfield_block[1])+idx)].value)
				if idx != int((len(self.groups[0])+1) * len(self.groups) -1)-1:
					self.getscore = self.getscore + '|'
					
		print("Returned Value are:")
		print(self.getscore)
		return self.getscore

if __name__ == '__main__':
	class_pytrackscore = pytrackscore()
